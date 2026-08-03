from __future__ import annotations

import logging
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Callable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

if TYPE_CHECKING:
    from doc_translator.translation import ModelTranslator

logger = logging.getLogger(__name__)


@dataclass
class _CellText:
    cell: Cell

    @property
    def text(self) -> str:
        return self.cell.value

    @property
    def translatable_text(self) -> str:
        return self.text.strip()

    def replace(self, text: str) -> None:
        original = self.text
        leading_length = len(original) - len(original.lstrip())
        trailing_length = len(original) - len(original.rstrip())
        leading = original[:leading_length]
        trailing = original[len(original) - trailing_length :] if trailing_length else ""
        self.cell.value = f"{leading}{text}{trailing}"
        self.cell.data_type = "s"


def translate_xlsx(
    input_path: str,
    output_path: Path,
    *,
    translator: ModelTranslator,
    translate_segments: Callable[..., list[str]],
    source_language: str,
    target_language: str,
    on_progress: Callable[[int, int], None],
    cancel_check: Callable[[], None],
    on_rebuilding: Callable[[], None],
) -> int | None:
    started_at = time.perf_counter()
    workbook = load_workbook(input_path, data_only=False, keep_links=True, rich_text=True)
    cells: list[_CellText] = []

    try:
        for worksheet in workbook.worksheets:
            cancel_check()
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or not isinstance(cell.value, str) or not cell.value.strip():
                        continue
                    cells.append(_CellText(cell))

        if cells:
            translated = translate_segments(
                translator,
                [cell.translatable_text for cell in cells],
                source_language=source_language,
                target_language=target_language,
                preserve_line_breaks=True,
                on_progress=on_progress,
                cancel_check=cancel_check,
            )
        else:
            translated = []
            logger.info("XLSX contains no translatable text", extra={"input": input_path})

        if len(translated) != len(cells):
            raise RuntimeError(f"XLSX translation returned {len(translated)} results for {len(cells)} text cells")

        for cell, translated_text in zip(cells, translated, strict=True):
            cell.replace(translated_text)

        cancel_check()
        on_rebuilding()
        temp_path: Path | None = None
        save_started_at = time.perf_counter()
        try:
            with tempfile.NamedTemporaryFile(
                dir=output_path.parent,
                prefix=f".{output_path.stem}.",
                suffix=".xlsx",
                delete=False,
            ) as temp_file:
                temp_path = Path(temp_file.name)
            workbook.save(temp_path)
            validation_workbook = load_workbook(temp_path, read_only=True, data_only=False, keep_links=True)
            validation_workbook.close()
            cancel_check()
            temp_path.replace(output_path)
        finally:
            if temp_path is not None:
                temp_path.unlink(missing_ok=True)

        logger.info(
            "Saved translated XLSX",
            extra={
                "output": str(output_path),
                "worksheets": len(workbook.worksheets),
                "cells": len(cells),
                "save_seconds": round(time.perf_counter() - save_started_at, 3),
                "total_seconds": round(time.perf_counter() - started_at, 3),
            },
        )
        return None
    finally:
        workbook.close()
